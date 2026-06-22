from flask import Flask, render_template_string, jsonify, request, session, redirect, url_for
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
import os
import re
import json

app = Flask(__name__)
# Chave de segurança para a tela de login (Sessões)
app.secret_key = "goiasa_roteirizacao_secreta"

# CONFIGURAÇÕES GERAIS
PASTA_PROJETO = os.path.dirname(os.path.abspath(__file__))
PLANILHA_PATH = os.path.join(PASTA_PROJETO, "Transporte Gestores (1).xlsx")
DB_ESTADO_PATH = os.path.join(PASTA_PROJETO, "estado_rotas.json")

# COORDENADA IMUTÁVEL DA SEDE GOIASA
SEDE_COORDS = [-18.07136465718134, -49.67222682214989]

# ==========================================
# 1. TEMPLATES DE INTERFACE (FRONTEND)
# ==========================================

LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Login - Roteirização Goiasa</title>
    <script src="https://cdn.tailwindcss.com"></script>
</head>
<body class="bg-slate-950 h-screen flex items-center justify-center relative overflow-hidden">
    <div class="absolute top-[-10%] left-[-10%] w-96 h-96 bg-emerald-600/20 rounded-full blur-[100px]"></div>
    <div class="absolute bottom-[-10%] right-[-10%] w-96 h-96 bg-blue-600/20 rounded-full blur-[100px]"></div>

    <div class="bg-slate-900 border border-slate-700/50 p-8 rounded-2xl shadow-2xl w-[400px] z-10 flex flex-col items-center">
        <img src="{{ url_for('static', filename='brand_green.png') }}" alt="Logo Goiasa" class="h-16 mb-6 object-contain">
        
        <h2 class="text-white text-xl font-bold mb-2 tracking-wide">Painel de Roteirização</h2>
        <p class="text-slate-400 text-sm mb-6">Insira suas credenciais para acessar</p>

        {% if erro %}
        <div class="bg-red-900/50 border border-red-500 text-red-300 text-xs px-4 py-2 rounded mb-4 w-full text-center">
            {{ erro }}
        </div>
        {% endif %}

        <form action="/login" method="POST" class="w-full">
            <div class="mb-4">
                <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Usuário</label>
                <input type="text" name="usuario" class="w-full bg-slate-950 border border-slate-700 rounded-lg p-3 text-white text-sm focus:border-emerald-500 outline-none transition-colors" placeholder="Ex: admin" required>
            </div>
            <div class="mb-6">
                <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Senha</label>
                <input type="password" name="senha" class="w-full bg-slate-950 border border-slate-700 rounded-lg p-3 text-white text-sm focus:border-emerald-500 outline-none transition-colors" placeholder="••••••••" required>
            </div>
            <button type="submit" class="w-full bg-emerald-600 hover:bg-emerald-500 text-white font-bold py-3 rounded-lg shadow-lg transition-colors flex items-center justify-center gap-2">
                🔒 Acessar Sistema
            </button>
        </form>
    </div>
</body>
</html>
"""

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Roteirização - Goiasa</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <style>
        body, html { height: 100vh; width: 100vw; margin: 0; padding: 0; overflow: hidden; font-family: ui-sans-serif, system-ui, sans-serif; }
        .grid-principal { display: grid; grid-template-columns: 78% 22%; height: calc(100vh - 3.5rem); width: 100vw; }
        .coluna-esquerda { display: grid; grid-template-rows: 70% 30%; height: 100%; }
        ::-webkit-scrollbar { width: 6px; height: 6px; }
        ::-webkit-scrollbar-track { background: #1e293b; }
        ::-webkit-scrollbar-thumb { background: #475569; border-radius: 3px; }
        ::-webkit-scrollbar-thumb:hover { background: #3b82f6; }
        .drag-over { border-color: #3b82f6 !important; background-color: #0f172a !important; }
        .leaflet-div-icon { background: transparent; border: none; }
    </style>
</head>
<body class="bg-slate-950 text-slate-100 relative">

    <div id="modalEdicao" class="hidden fixed inset-0 bg-black/80 z-[3000] flex items-center justify-center backdrop-blur-sm">
        <div class="bg-slate-800 p-6 rounded-xl w-[400px] border border-slate-600 shadow-2xl">
            <h3 class="text-white text-lg font-bold mb-4 flex items-center gap-2">✏️ Editar no Excel</h3>
            <input type="hidden" id="edit-id">
            <input type="hidden" id="edit-linha">
            <div class="mb-3">
                <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Nome</label>
                <input type="text" id="edit-nome" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-blue-500 outline-none">
            </div>
            <div class="mb-3 flex gap-3">
                <div class="flex-1">
                    <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Latitude</label>
                    <input type="text" id="edit-lat" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-blue-500 outline-none">
                </div>
                <div class="flex-1">
                    <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Longitude</label>
                    <input type="text" id="edit-lng" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-blue-500 outline-none">
                </div>
            </div>
            <div class="flex justify-end gap-3 mt-5">
                <button onclick="fecharModal('modalEdicao')" class="px-4 py-2 bg-slate-700 hover:bg-slate-600 text-white rounded text-sm transition-colors">Cancelar</button>
                <button onclick="salvarEdicao()" class="px-4 py-2 bg-blue-600 hover:bg-blue-500 text-white rounded font-bold text-sm shadow-lg transition-colors">💾 Salvar</button>
            </div>
        </div>
    </div>

    <div id="modalAdicionar" class="hidden fixed inset-0 bg-black/80 z-[3000] flex items-center justify-center backdrop-blur-sm">
        <div class="bg-slate-800 p-6 rounded-xl w-[450px] border border-emerald-600/50 shadow-2xl">
            <h3 class="text-white text-lg font-bold mb-4 flex items-center gap-2 text-emerald-400">➕ Cadastrar Novo Passageiro</h3>
            <div class="mb-3">
                <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Nome Completo *</label>
                <input type="text" id="add-nome" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-emerald-500 outline-none">
            </div>
            <div class="mb-3 flex gap-3">
                <div class="flex-1">
                    <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Latitude *</label>
                    <input type="text" id="add-lat" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-emerald-500 outline-none">
                </div>
                <div class="flex-1">
                    <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Longitude *</label>
                    <input type="text" id="add-lng" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-emerald-500 outline-none">
                </div>
            </div>
            <div class="mb-3 flex gap-3">
                <div class="flex-[2]">
                    <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Endereço (Opcional)</label>
                    <input type="text" id="add-end" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-emerald-500 outline-none">
                </div>
                <div class="flex-1">
                    <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Bairro (Opcional)</label>
                    <input type="text" id="add-bairro" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-emerald-500 outline-none">
                </div>
            </div>
            <div class="flex justify-end gap-3 mt-5">
                <button onclick="fecharModal('modalAdicionar')" class="px-4 py-2 bg-slate-700 hover:bg-slate-600 text-white rounded text-sm transition-colors">Cancelar</button>
                <button onclick="salvarNovoPassageiro()" class="px-4 py-2 bg-emerald-600 hover:bg-emerald-500 text-white rounded font-bold text-sm shadow-lg transition-colors">➕ Cadastrar</button>
            </div>
        </div>
    </div>

    <div id="modalMotorista" class="hidden fixed inset-0 bg-black/80 z-[3000] flex items-center justify-center backdrop-blur-sm">
        <div class="bg-slate-800 p-6 rounded-xl w-[350px] border border-blue-600/50 shadow-2xl">
            <h3 class="text-white text-lg font-bold mb-4 flex items-center gap-2 text-blue-400">👨‍✈️ Editar Motorista</h3>
            <input type="hidden" id="mot-carro-id">
            <div class="mb-4">
                <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Nome do Motorista Titular</label>
                <input type="text" id="mot-nome" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-blue-500 outline-none" placeholder="Deixe em branco para remover">
            </div>
            <div class="flex justify-end gap-3">
                <button onclick="fecharModal('modalMotorista')" class="px-4 py-2 bg-slate-700 hover:bg-slate-600 text-white rounded text-sm transition-colors">Cancelar</button>
                <button onclick="salvarMotorista()" class="px-4 py-2 bg-blue-600 hover:bg-blue-500 text-white rounded font-bold text-sm shadow-lg transition-colors">Salvar</button>
            </div>
        </div>
    </div>

    <div id="modalCidadeCarro" class="hidden fixed inset-0 bg-black/80 z-[3000] flex items-center justify-center backdrop-blur-sm">
        <div class="bg-slate-800 p-6 rounded-xl w-[350px] border border-purple-600/50 shadow-2xl">
            <h3 class="text-white text-lg font-bold mb-4 flex items-center gap-2 text-purple-400">📍 Editar Cidade do Veículo</h3>
            <input type="hidden" id="cid-carro-id">
            <div class="mb-4">
                <label class="block text-slate-400 text-xs mb-1 uppercase tracking-wide">Selecione a Nova Base</label>
                <select id="cid-nome" class="w-full bg-slate-900 border border-slate-700 rounded p-2.5 text-white text-sm focus:border-purple-500 outline-none">
                    <option value="Itumbiara">Itumbiara</option>
                    <option value="Bom Jesus">Bom Jesus</option>
                    <option value="Goiatuba">Goiatuba</option>
                    <option value="Pontalina">Pontalina</option>
                </select>
            </div>
            <div class="flex justify-end gap-3">
                <button onclick="fecharModal('modalCidadeCarro')" class="px-4 py-2 bg-slate-700 hover:bg-slate-600 text-white rounded text-sm transition-colors">Cancelar</button>
                <button onclick="salvarCidadeCarro()" class="px-4 py-2 bg-purple-600 hover:bg-purple-500 text-white rounded font-bold text-sm shadow-lg transition-colors">Atualizar Cidade</button>
            </div>
        </div>
    </div>

    <header class="h-14 bg-slate-900 border-b border-slate-800 flex justify-between items-center px-6 shadow-md z-50 relative">
        <div class="flex items-center gap-3">
            <img src="{{ url_for('static', filename='brand_green.png') }}" alt="Goiasa Logo" class="h-8 object-contain">
            <h1 class="text-lg font-bold text-slate-200 tracking-wide border-l border-slate-700 pl-3">ROTEIRIZAÇÃO</h1>
        </div>
        
        <div class="flex items-center gap-4 bg-slate-800 px-4 py-1.5 rounded-lg border border-slate-700 shadow-inner">
            <a href="/?data={{ data_prev_iso }}" class="text-slate-400 hover:text-white font-bold text-lg px-2 cursor-pointer transition-colors">&lt;</a>
            <span class="font-mono text-sm text-slate-200 font-semibold tracking-widest" id="data-atual" data-iso="{{ data_atual_iso }}">{{ data_formatada }}</span>
            <a href="/?data={{ data_next_iso }}" class="text-slate-400 hover:text-white font-bold text-lg px-2 cursor-pointer transition-colors">&gt;</a>
        </div>
        
        <div class="flex items-center gap-2">
            <span class="text-xs text-slate-500 font-medium bg-slate-800 px-3 py-1.5 rounded-full mr-2">Sede: Goiatuba-GO</span>
            <button onclick="limparRoteirizacao()" class="bg-red-700 hover:bg-red-600 text-white text-[11px] font-bold py-1.5 px-3 rounded-lg shadow transition-colors flex items-center gap-1 border border-red-600">🗑️ Limpar</button>
            <button onclick="copiarRotaAnterior()" class="bg-slate-700 hover:bg-slate-600 text-slate-300 text-[11px] font-bold py-1.5 px-3 rounded-lg shadow transition-colors flex items-center gap-1 border border-slate-600">⏳ Repetir Rota</button>
            <button onclick="exportarExcel()" class="bg-emerald-700 hover:bg-emerald-600 text-white text-[11px] font-bold py-1.5 px-3 rounded-lg shadow transition-colors flex items-center gap-1 border border-emerald-600">📥 Exportar</button>
            <a href="/logout" class="ml-2 bg-slate-800 hover:bg-slate-700 text-slate-300 text-[11px] font-bold py-1.5 px-3 rounded-lg shadow transition-colors border border-slate-600" title="Sair">Sair</a>
        </div>
    </header>

    <div class="grid-principal">
        <div class="coluna-esquerda border-r border-slate-800">
            <div id="mapa" class="w-full h-full relative bg-slate-950/50">
                <button onclick="calcularRotaIdeal()" class="absolute top-4 right-4 z-[1000] bg-blue-600 hover:bg-blue-700 text-white font-bold py-2.5 px-5 rounded-lg shadow-xl transition-all border border-blue-500 cursor-pointer text-sm tracking-wide">
                    ✨ Rota Ideal
                </button>
            </div>

            <div class="bg-slate-900 p-4 flex flex-col h-full overflow-hidden">
                <div class="flex justify-between items-center mb-3">
                    <div>
                        <h2 class="text-xs font-bold text-slate-400 uppercase tracking-wider inline-block">Passageiros Disponíveis (<span id="count-disponiveis">0</span>)</h2>
                    </div>
                    <button onclick="abrirModalAdicionar()" class="bg-emerald-600 hover:bg-emerald-500 text-white text-[11px] font-bold py-1 px-3 rounded shadow transition-colors flex items-center gap-1">
                        ➕ Adicionar Novo
                    </button>
                </div>
                
                <div class="flex-1 overflow-x-auto flex gap-3 items-center pb-2" id="lista-passageiros">
                    {% for p in passageiros %}
                    {% if not p.carro_atual %}
                    <div id="pass-{{ p.id }}" 
                         draggable="{{ 'true' if p.atribuivel else 'false' }}" 
                         {% if p.atribuivel %} ondragstart="drag(event)" {% endif %}
                         data-id="{{ p.id }}"
                         data-nome="{{ p.nome }}" 
                         class="bg-slate-800 border {% if p.atribuivel %} border-slate-700 cursor-grab active:cursor-grabbing hover:border-red-500 {% else %} border-red-900/50 opacity-80 cursor-not-allowed {% endif %} rounded-lg p-3 min-w-[280px] max-w-[280px] h-[85px] shadow-md transition-all flex flex-col justify-between">
                        
                        <div class="flex justify-between items-start w-full gap-2">
                            <div class="overflow-hidden flex-1">
                                <p class="font-semibold text-xs truncate {% if not p.atribuivel %} text-red-400 {% else %} text-slate-200 {% endif %}" title="{{ p.nome }}">{{ p.nome }}</p>
                                <p class="text-[10px] text-slate-400 truncate mt-0.5" title="{{ p.endereco }}">{{ p.endereco }}</p>
                            </div>
                            <button onmousedown="event.stopPropagation()" onclick="abrirModalEdicao('{{ p.id }}', '{{ p.linha_excel }}', '{{ p.nome }}', '{{ p.lat if p.lat else '' }}', '{{ p.lng if p.lng else '' }}')" class="bg-slate-700 hover:bg-amber-600 text-white rounded p-1.5 text-[10px] transition-colors shadow flex-shrink-0 cursor-pointer" title="Editar Coordenadas">✏️</button>
                        </div>
                        
                        <div class="flex justify-between items-center text-[10px] uppercase font-medium mt-1">
                            {% if p.atribuivel %}
                                <span class="text-slate-500 truncate max-w-[150px]">{{ p.bairro }}</span>
                                <span class="text-slate-500">CEP: {{ p.cep }}</span>
                            {% else %}
                                <span class="text-red-500 font-bold bg-red-950/40 px-2 py-0.5 rounded border border-red-900 w-full text-center tracking-wider">FALTA COORDENADA</span>
                            {% endif %}
                        </div>
                    </div>
                    {% endif %}
                    {% endfor %}
                </div>
            </div>
        </div>

        <div id="painel-direito" class="bg-slate-900 p-4 flex flex-col h-full overflow-hidden">
            <div class="flex justify-between items-center mb-3">
                <h2 class="text-xs font-bold text-slate-400 uppercase tracking-wider flex items-center justify-between">
                    <span>Frota Operacional</span>
                </h2>
                <select id="filtroCidade" onchange="filtrarCarros()" class="bg-slate-800 text-slate-200 text-xs border border-slate-700 rounded p-1 outline-none focus:border-blue-500">
                    <option value="Todas">Todas as Cidades</option>
                    <option value="Itumbiara">Itumbiara</option>
                    <option value="Bom Jesus">Bom Jesus</option>
                    <option value="Goiatuba">Goiatuba</option>
                    <option value="Pontalina">Pontalina</option>
                </select>
            </div>

            <div class="space-y-3 flex-1 overflow-y-auto pr-1" id="lista-carros">
                {% for carro in frota %}
                {% set em_carro = passageiros | selectattr("carro_atual", "equalto", carro.id) | list %}
                
                {% set tem_motorista = 1 if carro.cidade == 'Itumbiara' else 0 %}
                {% set nome_motorista = motoristas.get(carro.id, carro.motorista_padrao) if carro.cidade == 'Itumbiara' else 'Sem Motorista Fixo' %}
                
                {% set total_ocupado = tem_motorista + em_carro|length %}
                {% set perc = total_ocupado / carro.lotacao_max %}
                
                {% set cor_bg = 'bg-emerald-900 text-emerald-400 border-emerald-700' if perc < 0.6 else ('bg-amber-900 text-amber-400 border-amber-700' if perc < 1.0 else 'bg-red-900 text-red-400 border-red-700') %}

                <div id="carro-{{ carro.id }}" data-cidade="{{ carro.cidade }}" ondragover="allowDrop(event)" ondragleave="dragLeave(event)" ondrop="drop(event)" class="bg-slate-800 border-2 border-slate-700/70 rounded-xl p-3.5 transition-all dropzone flex flex-col justify-between min-h-[140px] item-carro">
                    <div>
                        <div class="flex justify-between items-start mb-1">
                            <div class="flex items-center gap-1.5">
                                <span class="text-[11px] font-bold px-2 py-0.5 rounded text-white" style="background-color: {{ carro.cor }}">{{ carro.px }}</span>
                                <span class="text-xs font-mono text-slate-400 bg-slate-900 px-1.5 py-0.5 rounded border border-slate-700">{{ carro.modelo }}</span>
                                <span class="text-[9px] text-slate-500 uppercase tracking-wide ml-1 flex items-center gap-1 bg-slate-900 px-1.5 py-0.5 rounded cursor-pointer hover:bg-slate-700 transition" onclick="abrirModalCidade('{{ carro.id }}', '{{ carro.cidade }}')" title="Mudar Cidade do Carro">
                                    {{ carro.cidade }} ⚙️
                                </span>
                            </div>
                            <span id="badge-lota-{{ carro.id }}" data-tem-mot="{{ tem_motorista }}" class="text-[11px] font-bold px-2 py-0.5 rounded-full border {{ cor_bg }} transition-colors shadow-sm">
                                {{ total_ocupado }} / {{ carro.lotacao_max }}
                            </span>
                        </div>
                        
                        <div class="flex justify-between items-center text-[11px] mt-2 mb-2 bg-slate-900/50 rounded px-2 py-1 border border-slate-700/50">
                            <div class="flex items-center gap-1.5 truncate">
                                <span class="text-slate-500">👨‍✈️</span>
                                <span class="font-semibold {{ 'text-blue-400' if carro.cidade == 'Itumbiara' else 'text-slate-600' }} truncate">{{ nome_motorista }}</span>
                                {% if carro.cidade == 'Itumbiara' %}
                                <button onclick="abrirModalMotorista('{{ carro.id }}', '{{ motoristas.get(carro.id, '') }}')" class="ml-1 text-slate-400 hover:text-white transition-colors cursor-pointer" title="Editar Motorista">✏️</button>
                                {% endif %}
                            </div>
                            <span class="font-mono text-slate-400 font-semibold" id="km-{{ carro.id }}">0.00 KM</span>
                        </div>
                    </div>
                    
                    <div class="space-y-1.5 min-h-[50px] bg-slate-900/40 rounded-lg p-2 border border-dashed border-slate-700 text-xs flex flex-col justify-center" id="inside-{{ carro.id }}">
                        <p class="text-[10px] text-slate-600 text-center font-medium my-auto painel-vazio" style="{% if em_carro %}display:none;{% endif %}">Arraste passageiros para cá</p>
                        
                        {% for p in em_carro %}
                        <div class="mini-card-passageiro bg-slate-900/80 text-slate-200 p-1.5 rounded-md flex justify-between items-center text-[11px] border border-slate-700/60 shadow-inner" data-id="{{ p.id }}" data-nome="{{ p.nome }}">
                            <span class="truncate font-medium pr-1"><b class="text-slate-500 mr-1">ID:{{ p.id }}</b>{{ p.nome }}</span>
                            <button onclick="removerPassageiroDaMemoria('{{ p.id }}', '{{ carro.id }}', this)" class="text-red-400 hover:text-red-500 font-bold text-xs px-1.5 cursor-pointer">×</button>
                        </div>
                        {% endfor %}
                    </div>
                </div>
                {% endfor %}
            </div>
        </div>
    </div>

    <script>
        const SEDE_COORDS = {{ sede_coords | tojson }};
        const DATA_ATUAL = document.getElementById('data-atual').getAttribute('data-iso');
        
        var coresCarros = {
            {% for c in frota %}
            "{{ c.id }}": "{{ c.cor }}"{% if not loop.last %},{% endif %}
            {% endfor %}
        };

        var mapa = L.map('mapa', { zoomControl: false }).setView(SEDE_COORDS, 11);
        L.control.zoom({ position: 'topleft' }).addTo(mapa);
        L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', { attribution: '© OpenStreetMap' }).addTo(mapa);

        var sedeIcon = L.divIcon({
            html: `<div class="bg-red-600 w-7 h-7 rounded-full flex items-center justify-center border-2 border-white shadow-lg text-white font-bold text-xs">🏢</div>`,
            className: '', iconSize: [28, 28], iconAnchor: [14, 14]
        });
        L.marker(SEDE_COORDS, { icon: sedeIcon }).addTo(mapa).bindPopup(`<b>Sede Principal (Goiasa)</b>`).openPopup();

        var passageirosDados = {{ passageiros | tojson }};
        var rotasCarros = {}; 
        var sequenciaCarros = {}; 

        {% for carro in frota %}
            {% set em_carro = passageiros | selectattr("carro_atual", "equalto", carro.id) | list %}
            sequenciaCarros["{{ carro.id }}"] = {{ em_carro | map(attribute='id') | list | tojson }};
            rotasCarros["{{ carro.id }}"] = L.polyline([], { color: "{{ carro.cor }}", weight: 4, opacity: 0.85 }).addTo(mapa);
        {% endfor %}

        // FILTRO DE CIDADES
        function filtrarCarros() {
            let cidadeSelecionada = document.getElementById('filtroCidade').value;
            let carros = document.querySelectorAll('.item-carro');
            carros.forEach(carro => {
                if(cidadeSelecionada === 'Todas' || carro.getAttribute('data-cidade') === cidadeSelecionada) {
                    carro.style.display = 'flex';
                } else {
                    carro.style.display = 'none';
                }
            });
        }

        // FUNÇÕES DE ABERTURA DE MODAIS
        function abrirModalEdicao(id, linha, nome, lat, lng) {
            document.getElementById('edit-id').value = id;
            document.getElementById('edit-linha').value = linha;
            document.getElementById('edit-nome').value = nome;
            document.getElementById('edit-lat').value = lat;
            document.getElementById('edit-lng').value = lng;
            document.getElementById('modalEdicao').classList.remove('hidden');
        }

        function abrirModalAdicionar() {
            document.getElementById('add-nome').value = '';
            document.getElementById('add-lat').value = '';
            document.getElementById('add-lng').value = '';
            document.getElementById('add-end').value = '';
            document.getElementById('add-bairro').value = '';
            document.getElementById('modalAdicionar').classList.remove('hidden');
        }

        function abrirModalMotorista(carId, nomeAtual) {
            document.getElementById('mot-carro-id').value = carId;
            document.getElementById('mot-nome').value = nomeAtual;
            document.getElementById('modalMotorista').classList.remove('hidden');
        }

        function abrirModalCidade(carId, cidadeAtual) {
            document.getElementById('cid-carro-id').value = carId;
            document.getElementById('cid-nome').value = cidadeAtual;
            document.getElementById('modalCidadeCarro').classList.remove('hidden');
        }

        function fecharModal(idModal) { document.getElementById(idModal).classList.add('hidden'); }

        // AÇÕES DA FERRAMENTA
        function limparRoteirizacao() {
            if(confirm("⚠️ ATENÇÃO: Tem certeza que deseja LIMPAR TODA A ROTEIRIZAÇÃO do dia " + document.getElementById('data-atual').innerText + "?")) {
                fetch('/api/limpar-roteirizacao', {
                    method: 'POST', headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ data: DATA_ATUAL })
                }).then(res => res.json()).then(data => {
                    if(data.status === 'sucesso') window.location.reload();
                });
            }
        }

        function copiarRotaAnterior() { alert("Aguarde! O recurso [Repetir Rota] será disponibilizado na próxima atualização."); }

        function exportarExcel() {
            fetch(`/api/exportar?data=${DATA_ATUAL}`)
            .then(res => res.json())
            .then(data => {
                if(data.status === 'sucesso') alert("✅ Relatório Exportado com Sucesso!\\nArquivo salvo como:\\n" + data.arquivo);
                else alert("⛔ ERRO AO EXPORTAR:\\n" + data.mensagem);
            });
        }

        // REQUISIÇÕES DE SALVAMENTO
        function salvarEdicao() {
            let linha = document.getElementById('edit-linha').value;
            let nome = document.getElementById('edit-nome').value;
            let lat = document.getElementById('edit-lat').value;
            let lng = document.getElementById('edit-lng').value;
            fetch('/api/editar-passageiro', {
                method: 'POST', headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ linha_excel: linha, nome: nome, lat: lat, lng: lng })
            }).then(res => res.json()).then(data => {
                if(data.status === 'sucesso') window.location.reload(); 
                else alert("⛔ FECHE O EXCEL!\\n\\n" + data.mensagem);
            });
        }

        function salvarNovoPassageiro() {
            let nome = document.getElementById('add-nome').value.trim();
            let lat = document.getElementById('add-lat').value.trim();
            let lng = document.getElementById('add-lng').value.trim();
            let endereco = document.getElementById('add-end').value.trim();
            let bairro = document.getElementById('add-bairro').value.trim();
            if(!nome || !lat || !lng) { alert("Nome, Latitude e Longitude são obrigatórios!"); return; }
            fetch('/api/adicionar-passageiro', {
                method: 'POST', headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ nome: nome, lat: lat, lng: lng, endereco: endereco, bairro: bairro })
            }).then(res => res.json()).then(data => {
                if(data.status === 'sucesso') window.location.reload(); 
                else alert("⛔ FECHE O EXCEL ANTES DE SALVAR!\\n\\n" + data.mensagem);
            });
        }

        function salvarMotorista() {
            let carId = document.getElementById('mot-carro-id').value;
            let nome = document.getElementById('mot-nome').value.trim();
            fetch('/api/editar-motorista', {
                method: 'POST', headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ car_id: carId, nome: nome })
            }).then(res => res.json()).then(data => {
                if(data.status === 'sucesso') window.location.reload();
            });
        }

        function salvarCidadeCarro() {
            let carId = document.getElementById('cid-carro-id').value;
            let cidade = document.getElementById('cid-nome').value;
            fetch('/api/editar-cidade-carro', {
                method: 'POST', headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ car_id: carId, cidade: cidade })
            }).then(res => res.json()).then(data => {
                if(data.status === 'sucesso') window.location.reload();
            });
        }

        function notificarBackend() {
            let estado = {};
            document.querySelectorAll('.dropzone').forEach(car => {
                let carId = car.id.replace('carro-', '');
                let nomes = [];
                car.querySelectorAll('.mini-card-passageiro').forEach(el => nomes.push(el.getAttribute('data-nome')));
                estado[carId] = nomes;
            });
            fetch('/api/salvar-estado', { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify({ data: DATA_ATUAL, estado: estado }) });
        }

        // LÓGICA DE DRAG & DROP E PLOTAGEM DO MAPA MANTIDAS IGUAIS (APENAS OMITIDAS PARA RESUMO, O CÓDIGO FUNCIONA IGUAL A VERSÃO ANTERIOR)
        function atualizarContadorDisponiveis() {
            let count = 0;
            document.querySelectorAll('#lista-passageiros > div').forEach(el => {
                if(el.style.display !== 'none' && el.getAttribute('draggable') === 'true') count++;
            });
            document.getElementById('count-disponiveis').innerText = count;
        }

        function atualizarContadorVeh(carId, passageirosNoCarro, max) {
            let badge = document.getElementById(`badge-lota-${carId}`);
            let temMotorista = parseInt(badge.getAttribute('data-tem-mot'));
            let total_ocupado = temMotorista + passageirosNoCarro;
            
            badge.innerText = `${total_ocupado} / ${max}`;
            let perc = total_ocupado / max;
            badge.className = "text-[11px] font-bold px-2 py-0.5 rounded-full border transition-colors shadow-sm ";
            if (perc >= 1.0) badge.className += "bg-red-900 text-red-400 border-red-700";
            else if (perc >= 0.6) badge.className += "bg-amber-900 text-amber-400 border-amber-700";
            else badge.className += "bg-emerald-900 text-emerald-400 border-emerald-700";
        }

        passageirosDados.forEach((p) => {
            if(p.atribuivel && p.lat && p.lng) {
                let coords = [p.lat, p.lng];
                let isAlocado = p.carro_atual ? true : false;
                
                let primeiroNome = p.nome.split(' ')[0];
                let corFundo = isAlocado ? coresCarros[p.carro_atual] : '#dc2626'; // Vermelho
                let opacidade = isAlocado ? '0.9' : '1';
                let classExtra = isAlocado ? "pointer-events-none" : "hover:bg-red-500 cursor-grab active:cursor-grabbing";

                let passIcon = L.divIcon({
                    html: `<div id="marker-icon-${p.id}" draggable="${!isAlocado}" onmousedown="event.stopPropagation()" data-pass-id="${p.id}" style="background-color: ${corFundo}; opacity: ${opacidade};" class="${classExtra} text-white px-2 py-1 rounded-full flex items-center justify-center font-bold text-[10px] shadow-md border border-white transition-all whitespace-nowrap min-w-[30px]">${primeiroNome}</div>`,
                    className: '', iconSize: null
                });
                
                let marker = L.marker(coords, { icon: passIcon }).addTo(mapa);
                marker.bindPopup(`<b>${p.nome}</b><br>${p.endereco}`);

                if(!isAlocado) {
                    setTimeout(() => {
                        let el = document.getElementById(`marker-icon-${p.id}`);
                        if(el) { el.addEventListener('dragstart', (ev) => { ev.dataTransfer.setData("text/plain", `pass-${p.id}`); }); }
                    }, 300);
                }
            }
        });

        // DRAG & DROP EVENTOS
        function allowDrop(ev) { ev.preventDefault(); ev.currentTarget.classList.add('drag-over'); }
        function dragLeave(ev) { ev.currentTarget.classList.remove('drag-over'); }
        function drag(ev) { ev.dataTransfer.setData("text/plain", ev.target.id); }

        function drop(ev) {
            ev.preventDefault();
            let targetCar = ev.currentTarget;
            targetCar.classList.remove('drag-over');
            let data = ev.dataTransfer.getData("text");
            
            let passId = data.replace('pass-', '').replace('marker-icon-', '');
            let passengerEl = document.getElementById('pass-' + passId);
            if (!passengerEl || passengerEl.getAttribute('draggable') === 'false') return;

            let carId = targetCar.id.replace('carro-', '');
            let containerPassageiros = targetCar.querySelector('div[id^="inside-"]');
            
            let badge = document.getElementById(`badge-lota-${carId}`);
            let temMot = parseInt(badge.getAttribute('data-tem-mot'));
            let max = parseInt(badge.innerText.split('/')[1].trim());
            let ocupantesAtuais = containerPassageiros.querySelectorAll('.mini-card-passageiro').length;

            if((temMot + ocupantesAtuais + 1) > max) { alert("Capacidade de lotação esgotada!"); return; }

            let aviso = containerPassageiros.querySelector('.painel-vazio');
            if (aviso) aviso.style.display = 'none';
            
            sequenciaCarros[carId].push(passId);

            let nome = passengerEl.getAttribute('data-nome');
            let miniCard = document.createElement('div');
            miniCard.className = 'mini-card-passageiro bg-slate-900/80 text-slate-200 p-1.5 rounded-md flex justify-between items-center text-[11px] border border-slate-700/60 shadow-inner';
            miniCard.setAttribute('data-id', passId);
            miniCard.setAttribute('data-nome', nome);
            miniCard.innerHTML = `<span class="truncate font-medium pr-1"><b class="text-slate-500 mr-1">ID:${passId}</b>${nome}</span>
                                  <button onclick="removerPassageiroDaMemoria('${passId}', '${carId}', this)" class="text-red-400 hover:text-red-500 font-bold text-xs px-1.5 cursor-pointer">×</button>`;
            
            containerPassageiros.appendChild(miniCard);
            passengerEl.style.display = 'none';
            
            let iconEl = document.getElementById(`marker-icon-${passId}`);
            if(iconEl) {
                iconEl.style.backgroundColor = coresCarros[carId];
                iconEl.style.opacity = "0.9";
                iconEl.className = "text-white px-2 py-1 rounded-full flex items-center justify-center font-bold text-[10px] shadow-md border border-white pointer-events-none whitespace-nowrap min-w-[30px]";
                iconEl.setAttribute('draggable', 'false');
            }

            atualizarContadorVeh(carId, ocupantesAtuais + 1, max);
            recalcularLinhaRota(carId);
            atualizarContadorDisponiveis();
            notificarBackend(); 
        }

        function removerPassageiroDaMemoria(passId, carId, botao) {
            sequenciaCarros[carId] = sequenciaCarros[carId].filter(id => id != passId);
            botao.parentElement.remove();
            
            let badge = document.getElementById(`badge-lota-${carId}`);
            let max = parseInt(badge.innerText.split('/')[1].trim());
            let containerPassageiros = document.getElementById(`inside-${carId}`);
            let ocupantesAtuais = containerPassageiros.querySelectorAll('.mini-card-passageiro').length;

            let passengerEl = document.getElementById(`pass-${passId}`);
            if(passengerEl) passengerEl.style.display = 'flex';
            else { notificarBackend(); window.location.reload(); return; }
            
            let iconEl = document.getElementById(`marker-icon-${passId}`);
            if(iconEl) {
                iconEl.style.backgroundColor = "#dc2626"; 
                iconEl.style.opacity = "1";
                iconEl.className = "text-white hover:bg-red-500 cursor-grab active:cursor-grabbing transition-all px-2 py-1 rounded-full flex items-center justify-center font-bold text-[10px] shadow-md border border-white whitespace-nowrap min-w-[30px]";
                iconEl.setAttribute('draggable', 'true');
            }

            if (ocupantesAtuais === 0) {
                let aviso = containerPassageiros.querySelector('.painel-vazio');
                if (aviso) aviso.style.display = 'block';
            }
            
            atualizarContadorVeh(carId, ocupantesAtuais, max);
            recalcularLinhaRota(carId);
            atualizarContadorDisponiveis();
            notificarBackend();
        }

        // MOTOR OSRM
        async function recalcularLinhaRota(carId) {
            let listPassIds = sequenciaCarros[carId];
            let arrayCoords = [SEDE_COORDS]; 

            listPassIds.forEach(id => {
                let p = passageirosDados.find(item => item.id == id);
                if(p && p.lat) arrayCoords.push([p.lat, p.lng]);
            });

            if (arrayCoords.length < 2) {
                rotasCarros[carId].setLatLngs([]); 
                document.getElementById(`km-${carId}`).innerText = `0.00 KM`;
                return;
            }

            let coordenadasOSRM = arrayCoords.map(c => `${c[1]},${c[0]}`).join(';');
            let url = `https://router.project-osrm.org/route/v1/driving/${coordenadasOSRM}?overview=full&geometries=geojson`;

            try {
                let response = await fetch(url);
                let data = await response.json();
                if (data.code === 'Ok') {
                    let distanciaReal = data.routes[0].distance;
                    document.getElementById(`km-${carId}`).innerText = `${(distanciaReal / 1000).toFixed(2)} KM`;
                    let coordenadasReais = data.routes[0].geometry.coordinates.map(c => [c[1], c[0]]);
                    rotasCarros[carId].setLatLngs(coordenadasReais);
                } else fallbackLinhaReta(carId, arrayCoords);
            } catch (erro) {
                fallbackLinhaReta(carId, arrayCoords);
            }
        }

        function fallbackLinhaReta(carId, arrayCoords) {
            rotasCarros[carId].setLatLngs(arrayCoords);
            let acumuladoMetros = 0;
            for (let i = 0; i < arrayCoords.length - 1; i++) {
                let pontoA = L.latLng(arrayCoords[i]);
                let pontoB = L.latLng(arrayCoords[i+1]);
                acumuladoMetros += pontoA.distanceTo(pontoB);
            }
            document.getElementById(`km-${carId}`).innerText = `${(acumuladoMetros / 1000).toFixed(2)} KM`;
        }

        function calcularRotaIdeal() { alert("Aguardando motor de otimização..."); }
        Object.keys(sequenciaCarros).forEach(carId => { recalcularLinhaRota(carId); });
    </script>
</body>
</html>
"""

# ==========================================
# 2. SISTEMA DE BACKEND E EXCEL
# ==========================================

# Middleware: Exige Login em todas as rotas
@app.before_request
def require_login():
    endpoints_publicos = ['login', 'static']
    if request.endpoint not in endpoints_publicos and 'logged_in' not in session:
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        usuario = request.form.get('usuario')
        senha = request.form.get('senha')
        
        # Acesso Padrão 
        if usuario == 'admin' and senha == 'admin':
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            erro = "Usuário ou senha incorretos. Tente novamente."
            
    return render_template_string(LOGIN_TEMPLATE, erro=erro)

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))


def gerenciar_banco_estado():
    if not os.path.exists(DB_ESTADO_PATH):
        with open(DB_ESTADO_PATH, 'w', encoding='utf-8') as f:
            json.dump({"motoristas": {}, "cidades_carros": {}}, f)
        return {"motoristas": {}, "cidades_carros": {}}
    with open(DB_ESTADO_PATH, 'r', encoding='utf-8') as f:
        db = json.load(f)
        if "motoristas" not in db: db["motoristas"] = {}
        if "cidades_carros" not in db: db["cidades_carros"] = {}
        return db

def limpar_coordenada(valor):
    if pd.isna(valor) or str(valor).strip() == "" or str(valor).lower() == "nan": return None
    texto = str(valor).strip().replace(',', '.')
    match = re.search(r'-?\d+\.\d+', texto)
    if match: return float(match.group())
    return None

def carregar_passageiros_excel(data_atual):
    passageiros = []
    if not os.path.exists(PLANILHA_PATH): return []
    
    estado_db = gerenciar_banco_estado()
    estado_hoje = estado_db.get(data_atual, {})
    
    pass_in_car = {}
    for carro_id, nomes in estado_hoje.items():
        for n in nomes: pass_in_car[n] = carro_id
    
    frota_valida = [c['id'] for c in gerar_frota_operacional()]

    try:
        df = pd.read_excel(PLANILHA_PATH, sheet_name='BASE', header=None, dtype=str)
        counter = 1
        
        for i in range(1, len(df)):
            linha_absoluta = i + 1 
            row = df.iloc[i]
            
            nome_val = str(row[1]) if len(row) > 1 and not pd.isna(row[1]) else ""
            if nome_val.strip() == "" or nome_val.lower() == "nan": continue

            lat_crua = row[6] if len(row) > 6 else None
            lng_crua = row[7] if len(row) > 7 else None
            lat_limpa = limpar_coordenada(lat_crua)
            lng_limpa = limpar_coordenada(lng_crua)
            
            endereco = str(row[2]).strip() if len(row) > 2 and not pd.isna(row[2]) else ""
            bairro = str(row[3]).strip() if len(row) > 3 and not pd.isna(row[3]) else ""
            cep = str(row[4]).strip() if len(row) > 4 and not pd.isna(row[4]) else ""
            
            atribuivel = False
            if lat_limpa is not None and lng_limpa is not None:
                atribuivel = True
            
            carro_atual = pass_in_car.get(nome_val.strip(), None)
            if carro_atual and carro_atual not in frota_valida:
                carro_atual = None
                
            passageiros.append({
                "id": counter,
                "linha_excel": linha_absoluta,
                "nome": nome_val.strip(),
                "endereco": endereco,
                "bairro": bairro,
                "cep": cep,
                "lat": lat_limpa,
                "lng": lng_limpa,
                "atribuivel": atribuivel,
                "carro_atual": carro_atual
            })
            counter += 1
    except Exception as e:
        print(f"Erro na leitura: {e}")
        
    return passageiros

def gerar_frota_operacional():
    # Pega as cidades salvas do banco de dados (que você editou)
    estado_db = gerenciar_banco_estado()
    cidades_salvas = estado_db.get("cidades_carros", {})

    frota_base = [
        {"id": "V_ITU_1685", "px": "PX-1685", "modelo": "Spin", "cidade_padrao": "Itumbiara", "lotacao_max": 7, "cor": "#3b82f6", "motorista_padrao": "Henrique/Feliciano"},
        {"id": "V_ITU_1722", "px": "PX-1722", "modelo": "Spin", "cidade_padrao": "Itumbiara", "lotacao_max": 7, "cor": "#10b981", "motorista_padrao": "Gilson/Carlos Augusto"},
        {"id": "V_ITU_1683", "px": "PX-1683", "modelo": "Spin", "cidade_padrao": "Itumbiara", "lotacao_max": 7, "cor": "#f59e0b", "motorista_padrao": "Paulo/Mirclênio"},
        {"id": "V_ITU_1687", "px": "PX-1687", "modelo": "Spin", "cidade_padrao": "Itumbiara", "lotacao_max": 7, "cor": "#ec4899", "motorista_padrao": "Walles/Marcos"},
        {"id": "V_ITU_1688", "px": "PX-1688", "modelo": "Spin", "cidade_padrao": "Itumbiara", "lotacao_max": 7, "cor": "#8b5cf6", "motorista_padrao": "Alaerte/Eugênio"},
        {"id": "V_ITU_1719", "px": "PX-1719", "modelo": "Spin", "cidade_padrao": "Itumbiara", "lotacao_max": 7, "cor": "#06b6d4", "motorista_padrao": "Luciana/José/Daniela"},

        {"id": "V_BJ_1670", "px": "PX-1670", "modelo": "Polo", "cidade_padrao": "Bom Jesus", "lotacao_max": 5, "cor": "#f43f5e", "motorista_padrao": ""},
        {"id": "V_BJ_1733", "px": "PX-1733", "modelo": "Polo", "cidade_padrao": "Bom Jesus", "lotacao_max": 5, "cor": "#eab308", "motorista_padrao": ""},
        {"id": "V_BJ_1738", "px": "PX-1738", "modelo": "Argo", "cidade_padrao": "Bom Jesus", "lotacao_max": 5, "cor": "#14b8a6", "motorista_padrao": ""},
        {"id": "V_BJ_1662", "px": "PX-1662", "modelo": "Polo", "cidade_padrao": "Bom Jesus", "lotacao_max": 5, "cor": "#84cc16", "motorista_padrao": ""},
        {"id": "V_BJ_1684", "px": "PX-1684", "modelo": "Spin", "cidade_padrao": "Bom Jesus", "lotacao_max": 7, "cor": "#6366f1", "motorista_padrao": ""},
        {"id": "V_BJ_1674", "px": "PX-1674", "modelo": "Polo", "cidade_padrao": "Bom Jesus", "lotacao_max": 5, "cor": "#a855f7", "motorista_padrao": ""},
        {"id": "V_BJ_1655", "px": "PX-1655", "modelo": "Polo", "cidade_padrao": "Bom Jesus", "lotacao_max": 5, "cor": "#d946ef", "motorista_padrao": ""},
        {"id": "V_BJ_1680", "px": "PX-1680", "modelo": "Polo", "cidade_padrao": "Bom Jesus", "lotacao_max": 5, "cor": "#ec4899", "motorista_padrao": ""},
        {"id": "V_BJ_1706", "px": "PX-1706", "modelo": "Oroch", "cidade_padrao": "Bom Jesus", "lotacao_max": 2, "cor": "#f97316", "motorista_padrao": ""},
        {"id": "V_BJ_1700", "px": "PX-1700", "modelo": "Strada", "cidade_padrao": "Bom Jesus", "lotacao_max": 2, "cor": "#0ea5e9", "motorista_padrao": ""},

        {"id": "V_GOI_1678", "px": "PX-1678", "modelo": "Polo", "cidade_padrao": "Goiatuba", "lotacao_max": 5, "cor": "#3b82f6", "motorista_padrao": ""},
        {"id": "V_GOI_1687_GOI", "px": "PX-1687", "modelo": "Spin", "cidade_padrao": "Goiatuba", "lotacao_max": 7, "cor": "#10b981", "motorista_padrao": ""},
        {"id": "V_GOI_1657", "px": "PX-1657", "modelo": "Polo", "cidade_padrao": "Goiatuba", "lotacao_max": 5, "cor": "#f59e0b", "motorista_padrao": ""},
        {"id": "V_GOI_1726", "px": "PX-1726", "modelo": "Strada", "cidade_padrao": "Goiatuba", "lotacao_max": 2, "cor": "#ef4444", "motorista_padrao": ""},
        {"id": "V_GOI_1730", "px": "PX-1730", "modelo": "Polo", "cidade_padrao": "Goiatuba", "lotacao_max": 5, "cor": "#8b5cf6", "motorista_padrao": ""},
        {"id": "V_GOI_1686", "px": "PX-1686", "modelo": "Spin", "cidade_padrao": "Goiatuba", "lotacao_max": 7, "cor": "#06b6d4", "motorista_padrao": ""},
        {"id": "V_GOI_1689", "px": "PX-1689", "modelo": "Spin", "cidade_padrao": "Goiatuba", "lotacao_max": 7, "cor": "#f43f5e", "motorista_padrao": ""},
        {"id": "V_GOI_1725", "px": "PX-1725", "modelo": "Spin", "cidade_padrao": "Goiatuba", "lotacao_max": 7, "cor": "#64748b", "motorista_padrao": ""},

        {"id": "V_PON_1723", "px": "PX-1723", "modelo": "Spin", "cidade_padrao": "Pontalina", "lotacao_max": 7, "cor": "#14b8a6", "motorista_padrao": ""},
    ]

    # Aplica a cidade salva se existir, senão usa a padrão
    for carro in frota_base:
        carro["cidade"] = cidades_salvas.get(carro["id"], carro["cidade_padrao"])
        
    return frota_base

# ----------------- ROTAS DO SERVIDOR -----------------

@app.route('/')
def index():
    data_req = request.args.get('data', '2026-06-21')
    try: data_obj = datetime.strptime(data_req, "%Y-%m-%d")
    except ValueError: data_obj = datetime.strptime('2026-06-21', "%Y-%m-%d")
        
    data_formatada = data_obj.strftime("%d/%m/%Y")
    data_prev_iso = (data_obj - timedelta(days=1)).strftime("%Y-%m-%d")
    data_next_iso = (data_obj + timedelta(days=1)).strftime("%Y-%m-%d")

    passageiros = carregar_passageiros_excel(data_req)
    frota = gerar_frota_operacional()
    
    estado_db = gerenciar_banco_estado()
    motoristas = estado_db.get("motoristas", {})
    
    return render_template_string(HTML_TEMPLATE, passageiros=passageiros, frota=frota, sede_coords=SEDE_COORDS, 
                                  data_atual_iso=data_req, data_formatada=data_formatada, 
                                  data_prev_iso=data_prev_iso, data_next_iso=data_next_iso, motoristas=motoristas)

@app.route('/api/editar-cidade-carro', methods=['POST'])
def editar_cidade_carro():
    dados = request.json
    car_id = dados.get('car_id')
    cidade_nova = dados.get('cidade')

    db = gerenciar_banco_estado()
    if 'cidades_carros' not in db: db['cidades_carros'] = {}
    
    db['cidades_carros'][car_id] = cidade_nova

    with open(DB_ESTADO_PATH, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=4)

    return jsonify({"status": "sucesso"})

@app.route('/api/editar-motorista', methods=['POST'])
def editar_motorista():
    dados = request.json
    car_id = dados.get('car_id')
    nome = dados.get('nome')
    db = gerenciar_banco_estado()
    if 'motoristas' not in db: db['motoristas'] = {}
    if not nome or str(nome).strip() == "":
        if car_id in db['motoristas']: del db['motoristas'][car_id]
    else:
        db['motoristas'][car_id] = str(nome).strip()
    with open(DB_ESTADO_PATH, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=4)
    return jsonify({"status": "sucesso"})

@app.route('/api/salvar-estado', methods=['POST'])
def salvar_estado():
    dados = request.json
    db = gerenciar_banco_estado()
    db[dados.get('data')] = dados.get('estado')
    with open(DB_ESTADO_PATH, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=4)
    return jsonify({"status": "sucesso"})

@app.route('/api/limpar-roteirizacao', methods=['POST'])
def limpar_roteirizacao():
    dados = request.json
    data_req = dados.get('data')
    db = gerenciar_banco_estado()
    if data_req in db:
        db[data_req] = {} 
        with open(DB_ESTADO_PATH, 'w', encoding='utf-8') as f:
            json.dump(db, f, ensure_ascii=False, indent=4)
    return jsonify({"status": "sucesso"})

@app.route('/api/editar-passageiro', methods=['POST'])
def editar_passageiro():
    dados = request.json
    try:
        linha = int(dados['linha_excel'])
        wb = openpyxl.load_workbook(PLANILHA_PATH)
        ws = wb['BASE']
        ws.cell(row=linha, column=2).value = str(dados['nome']).strip()
        ws.cell(row=linha, column=7).value = str(dados['lat']).strip().replace(',', '.')
        ws.cell(row=linha, column=8).value = str(dados['lng']).strip().replace(',', '.')
        wb.save(PLANILHA_PATH)
        return jsonify({"status": "sucesso"})
    except PermissionError: return jsonify({"status": "erro", "mensagem": "Arquivo em uso pelo Excel."})
    except Exception as e: return jsonify({"status": "erro", "mensagem": str(e)})

@app.route('/api/adicionar-passageiro', methods=['POST'])
def adicionar_passageiro():
    dados = request.json
    try:
        wb = openpyxl.load_workbook(PLANILHA_PATH)
        ws = wb['BASE']
        nova_linha = ws.max_row + 1
        ws.cell(row=nova_linha, column=2).value = str(dados['nome']).strip()
        ws.cell(row=nova_linha, column=3).value = str(dados.get('endereco', '')).strip()
        ws.cell(row=nova_linha, column=4).value = str(dados.get('bairro', '')).strip()
        ws.cell(row=nova_linha, column=7).value = str(dados['lat']).strip().replace(',', '.')
        ws.cell(row=nova_linha, column=8).value = str(dados['lng']).strip().replace(',', '.')
        wb.save(PLANILHA_PATH)
        return jsonify({"status": "sucesso"})
    except PermissionError: return jsonify({"status": "erro", "mensagem": "Arquivo em uso pelo Excel."})
    except Exception as e: return jsonify({"status": "erro", "mensagem": str(e)})

@app.route('/api/exportar')
def exportar_excel():
    data_req = request.args.get('data')
    if not data_req: return jsonify({"status": "erro", "mensagem": "Data não fornecida."})
    passageiros = carregar_passageiros_excel(data_req)
    frota = {c['id']: c for c in gerar_frota_operacional()}
    estado_db = gerenciar_banco_estado()
    motoristas = estado_db.get("motoristas", {})
    dados_export = []
    
    for p in passageiros:
        carro_id = p.get('carro_atual')
        if carro_id:
            veiculo = frota.get(carro_id, {})
            px = veiculo.get('px', '-')
            cidade = veiculo.get('cidade', '-')
            motorista = motoristas.get(carro_id, veiculo.get('motorista_padrao', 'Sem Motorista Fixo')) if veiculo.get('cidade') == 'Itumbiara' else 'N/A'
            dados_export.append({
                "ID": p['id'],
                "Passageiro": p['nome'],
                "Endereço": p['endereco'],
                "Bairro": p['bairro'],
                "Cidade Base do Carro": cidade,
                "Veículo (PX)": px,
                "Motorista": motorista
            })
        else:
            dados_export.append({
                "ID": p['id'],
                "Passageiro": p['nome'],
                "Endereço": p['endereco'],
                "Bairro": p['bairro'],
                "Cidade Base do Carro": "Não Alocado",
                "Veículo (PX)": "Não Alocado",
                "Motorista": "-"
            })
            
    df_export = pd.DataFrame(dados_export)
    df_export.sort_values(by=["Cidade Base do Carro", "Veículo (PX)", "Passageiro"], inplace=True)
    nome_arquivo = f"Roteirizacao_{data_req}.xlsx"
    caminho_salvar = os.path.join(PASTA_PROJETO, nome_arquivo)
    
    try:
        df_export.to_excel(caminho_salvar, index=False)
        return jsonify({"status": "sucesso", "arquivo": nome_arquivo})
    except PermissionError: return jsonify({"status": "erro", "mensagem": f"O Excel está bloqueando a gravação do arquivo {nome_arquivo}."})
    except Exception as e: return jsonify({"status": "erro", "mensagem": str(e)})

if __name__ == '__main__':
    app.run(debug=True, port=5000)